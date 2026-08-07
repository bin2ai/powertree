"""Tests for tree analytics (efficiency/loss/top consumers) and CSV export."""

import csv
import math
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

from powertree import api  # noqa: E402
from powertree.model.elements import (  # noqa: E402
    PowerTree, Project, Source, Converter, Load, LoadType,
)
from powertree.model.calc import solve_tree  # noqa: E402


def test_tree_metrics_simple_converter_chain():
    t = PowerTree("t")
    src = t.add_element(Source(v_min=10, v_typ=10, v_max=10))
    c = t.add_element(Converter(efficiency_pct=80, vout_min=5, vout_typ=5,
                                vout_max=5), parent_id=src.id)
    t.add_element(Load(load_type=LoadType.CURRENT, value_typ=1.0),
                  parent_id=c.id)
    m = api.tree_metrics(t)
    # loads get 5 W, source delivers 6.25 W, loss 1.25 W, eff 80 %
    assert math.isclose(m["p_loads_typ"], 5.0, rel_tol=1e-6)
    assert math.isclose(m["p_source_typ"], 6.25, rel_tol=1e-6)
    assert math.isclose(m["p_loss_typ"], 1.25, rel_tol=1e-6)
    assert math.isclose(m["efficiency_pct"], 80.0, abs_tol=0.1)
    assert m["top_consumers"][0]["pct_of_source"] == 80.0


def test_tree_metrics_demo_consistency():
    p = api.demo_project()
    tree = p.trees[0]
    r = solve_tree(tree)
    m = api.tree_metrics(tree, r)
    # energy conservation: loads + losses ≈ source output
    assert math.isclose(m["p_loads_typ"] + m["p_loss_typ"],
                        m["p_source_typ"], rel_tol=1e-6)
    assert 0 < m["efficiency_pct"] < 100
    assert len(m["top_consumers"]) == 5
    pcts = [c["pct_of_source"] for c in m["top_consumers"]]
    assert pcts == sorted(pcts, reverse=True)


def test_project_summary_includes_metrics():
    s = api.project_summary(api.demo_project())
    t0 = s["trees"][0]
    assert t0["efficiency_pct"] is not None
    assert t0["top_consumers"]


def test_export_csv_roundtrips():
    p = api.demo_project()
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "table.csv")
        api.export_csv(p, path)
        with open(path, newline="", encoding="utf-8-sig") as fh:
            rows = list(csv.DictReader(fh))
    trees = {r["tree"] for r in rows}
    assert "Zynq Carrier 12V" in trees and "Battery Backup" in trees
    total_elements = sum(len(t.elements) for t in p.trees)
    assert len(rows) == total_elements
    vdda = next(r for r in rows if r["name"] == "VDDA 1.8V")
    assert vdda["status"] == "VIOLATION"
    src = next(r for r in rows if r["name"] == "12V DC Input")
    assert abs(float(src["pct_of_source_typ"]) - 100.0) < 0.1


def test_export_csv_via_generic_export():
    p = api.demo_project()
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "one.csv")
        api.export(p, "csv", path, "Battery Backup")
        with open(path, newline="", encoding="utf-8-sig") as fh:
            rows = list(csv.DictReader(fh))
    assert {r["tree"] for r in rows} == {"Battery Backup"}
    assert len(rows) == 3


def test_rail_headroom_math():
    from powertree.model.elements import LimitType
    t = PowerTree("t")
    src = t.add_element(Source(v_min=10, v_typ=10, v_max=10,
                               limit_type=LimitType.POWER, limit_value=10.0))
    c = t.add_element(Converter(efficiency_pct=100, vout_min=5, vout_typ=5,
                                vout_max=5, limit_type=LimitType.CURRENT,
                                limit_value=2.0), parent_id=src.id)
    t.add_element(Load(load_type=LoadType.CURRENT, value_typ=1.0),
                  parent_id=c.id)
    rows = api.rail_headroom(t)
    assert len(rows) == 2
    conv = next(r for r in rows if r["kind"] == "converter")
    assert math.isclose(conv["used_worst"], 1.0, rel_tol=1e-6)
    assert math.isclose(conv["headroom"], 1.0, rel_tol=1e-6)      # 1 A left
    assert math.isclose(conv["extra_load_w"], 5.0, rel_tol=1e-6)  # at 5 V
    src_row = next(r for r in rows if r["kind"] == "source")
    assert math.isclose(src_row["used_worst"], 5.0, rel_tol=1e-6)
    assert src_row["used_pct"] == 50.0
    # sorted tightest-first
    assert rows[0]["headroom_pct"] <= rows[-1]["headroom_pct"]


def test_rail_headroom_demo_flags_core_buck_tight():
    p = api.demo_project()
    rows = api.rail_headroom(p.trees[0])
    tightest = rows[0]
    assert "1.0V Core Buck" in tightest["name"]
    assert tightest["headroom_pct"] < 10


def test_excel_states_sheet():
    from openpyxl import load_workbook
    from powertree.export.excel_export import export_excel_xlsx
    p = api.demo_project()
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "r.xlsx")
        export_excel_xlsx(p, path)
        wb = load_workbook(path)
    assert "States" in wb.sheetnames
    ws = wb["States"]
    labels = [ws.cell(row=r, column=2).value for r in range(2, 10)]
    assert "Base" in labels and "Low Power" in labels \
        and "Performance" in labels
    assert "Efficiency (%)" in [c.value for c in wb["Overview"][5]]


def test_duplicate_subtree():
    p = api.demo_project()
    tree = p.trees[0]
    conv = api.find_element(tree, "U12")           # 1.8V buck w/ children
    n_before = len(tree.elements)
    subtree_size = 1 + len(tree.descendants_of(conv.id))
    dup = tree.duplicate_subtree(conv.id)
    assert len(tree.elements) == n_before + subtree_size
    assert dup.name.endswith("(copy)")
    assert dup.parent_id == conv.parent_id
    dup_desc = tree.descendants_of(dup.id)
    assert len(dup_desc) == subtree_size - 1
    assert {d.id for d in dup_desc}.isdisjoint(
        {d.id for d in tree.descendants_of(conv.id)})
    solve_tree(tree)                               # still solves cleanly
    with pytest.raises(ValueError, match="one source"):
        tree.duplicate_subtree(tree.source.id)


def test_derating_policy_in_validate():
    p = api.demo_project()
    assert p.derating_pct == 80.0
    v = api.validate(p)
    msgs = [f["message"] for f in v["findings"]]
    assert any("derating policy" in m and "1.0V Core Buck" in m
               for m in msgs), "92%-used core buck must trip 80% derating"
    p.derating_pct = 0                             # disable
    v2 = api.validate(p)
    assert not any("derating policy" in f["message"] for f in v2["findings"])


def test_derating_roundtrip():
    from powertree.model import serialization
    p = api.demo_project()
    p.derating_pct = 70.0
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "d.ptproj")
        serialization.save_project(p, path)
        assert serialization.load_project(path).derating_pct == 70.0


def test_html_report_export():
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    os.environ.setdefault("QT_QPA_FONTDIR", r"C:\Windows\Fonts")
    p = api.demo_project()
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "share.html")
        api.export(p, "html", path)
        with open(path, encoding="utf-8") as fh:
            doc = fh.read()
    assert "data:image/png;base64," in doc          # embedded flowchart
    assert "ATTENTION" in doc                       # verdict from findings
    assert "Rail budget" in doc and "Operating states" in doc
    assert "VDDA" in doc                            # findings listed
    assert len(doc) > 100_000                       # genuinely self-contained


def test_converter_efficiency_curve_interpolation():
    c = Converter(eff_points=[[0.1, 80.0], [1.0, 90.0]])
    assert math.isclose(c.efficiency_at(0.05), 0.80)     # clamp low
    assert math.isclose(c.efficiency_at(2.0), 0.90)      # clamp high
    assert math.isclose(c.efficiency_at(0.55), 0.85)     # midpoint
    flat = Converter(efficiency_pct=92.0)
    assert math.isclose(flat.efficiency_at(123.0), 0.92)  # no curve = flat


def test_solver_uses_efficiency_curve():
    t = PowerTree("t")
    src = t.add_element(Source(v_min=10, v_typ=10, v_max=10))
    c = t.add_element(Converter(
        efficiency_pct=50.0,                # would give p_in = 10 W if used
        eff_points=[[0.5, 80.0], [1.5, 80.0]],
        vout_min=5, vout_typ=5, vout_max=5), parent_id=src.id)
    t.add_element(Load(load_type=LoadType.CURRENT, value_typ=1.0),
                  parent_id=c.id)
    r = solve_tree(t)
    # curve says 80 % at 1 A: p_in = 5/0.8 = 6.25 W (not 10 W flat)
    assert math.isclose(r.get(c.id, "typ").p_in, 6.25, rel_tol=1e-6)


def test_efficiency_curve_roundtrip():
    from powertree.model import serialization
    p = Project("x")
    t = p.new_tree("m")
    src = t.add_element(Source())
    t.add_element(Converter(eff_points=[[0.1, 85.0], [1.0, 92.5]]),
                  parent_id=src.id)
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "c.ptproj")
        serialization.save_project(p, path)
        p2 = serialization.load_project(path)
    c2 = next(e for e in p2.trees[0].elements.values()
              if e.kind == "converter")
    assert c2.eff_points == [[0.1, 85.0], [1.0, 92.5]]


def test_parts_list_aggregates():
    p = api.demo_project()
    parts = api.parts_list(p)
    numbers = {x["part_number"]: x for x in parts}
    assert numbers["XC7Z020-1CLG484"]["count"] == 9      # one per Zynq rail
    assert numbers["XC7Z020-1CLG484"]["refdes"] == "U1"
    assert numbers["MT41K256M16"]["count"] == 4          # 2 loads x 2 chips
    assert "U2, U3" == numbers["MT41K256M16"]["refdes"]


def test_excel_parts_sheet():
    from openpyxl import load_workbook
    from powertree.export.excel_export import export_excel_xlsx
    p = api.demo_project()
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "r.xlsx")
        export_excel_xlsx(p, path)
        wb = load_workbook(path)
    assert "Parts" in wb.sheetnames
    col_a = [wb["Parts"].cell(row=r, column=1).value for r in range(2, 12)]
    assert "XC7Z020-1CLG484" in col_a


def test_fmt_si_digits_knob():
    from powertree.model import calc
    old = calc.SI_DIGITS
    try:
        calc.SI_DIGITS = 5
        assert calc.fmt_si(1.23456, "W") == "1.2346 W"
        calc.SI_DIGITS = 3
        assert calc.fmt_si(1.23456, "W") == "1.23 W"
    finally:
        calc.SI_DIGITS = old


def test_waiver_lifecycle():
    from powertree.model import serialization
    from powertree.model.calc import solve_tree as _solve
    p = api.demo_project()
    tree = p.trees[0]
    r = _solve(tree)
    vdda_warns = [w for w in r.warnings if "VDDA" in w.message]
    assert len(vdda_warns) == 3
    with pytest.raises(ValueError, match="justification"):
        api.waive_finding(p, vdda_warns[0].element_id,
                          vdda_warns[0].message, "   ")
    for w in vdda_warns:
        api.waive_finding(p, w.element_id, w.message,
                          "Clock gen VDDA has internal LDO, tolerates 1.5 V "
                          "per vendor errata EN-042.")
    active, waived = api.split_waived(p, r.warnings)
    assert len(waived) == 3
    assert not any("VDDA" in w.message for w in active)
    v = api.validate(p)
    base_errors = [f for f in v["findings"]
                   if f["severity"] == "error" and f["state"] == "Base"
                   and not f["waived"]]
    assert base_errors == []          # all three errors waived
    assert v["waived"] >= 3
    # persists through the file format
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, "w.ptproj")
        serialization.save_project(p, path)
        p2 = serialization.load_project(path)
    assert len(p2.waivers) == 3
    # unwaive works
    assert api.unwaive_finding(p, vdda_warns[0].element_id,
                               vdda_warns[0].message)
    active2, waived2 = api.split_waived(p, r.warnings)
    assert len(waived2) == 2


def test_demo_uses_efficiency_curve():
    p = api.demo_project()
    tree = p.trees[0]
    buck5 = api.find_element(tree, "U10")
    assert len(buck5.eff_points) >= 5
    r = solve_tree(tree)
    i_out = r.get(buck5.id, "typ").i_out
    eff = buck5.efficiency_at(i_out)
    # at ~0.9 A the curve sits between the 0.5 A and 1.0 A points
    assert 0.90 < eff < 0.94


def test_export_bundle():
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    os.environ.setdefault("QT_QPA_FONTDIR", r"C:\Windows\Fonts")
    p = api.demo_project()
    with tempfile.TemporaryDirectory() as td:
        written = api.export_bundle(p, td)
        names = {os.path.basename(w) for w in written}
        assert len(written) == 6      # pdf, html, xlsx, csv, 2 tree PNGs
        assert any(n.endswith("_report.pdf") for n in names)
        assert any(n.endswith("_report.html") for n in names)
        assert any(n.endswith("_report.xlsx") for n in names)
        assert any(n.endswith("_table.csv") for n in names)
        assert sum(1 for n in names if n.endswith(".png")) == 2
        for w in written:
            assert os.path.getsize(w) > 500


def test_cli_validate_strict(demo_path=None):
    import subprocess
    root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    env = dict(os.environ, PYTHONPATH=os.path.join(root, "src"))
    with tempfile.TemporaryDirectory() as td:
        clean = os.path.join(td, "clean.ptproj")
        p = Project("clean")
        t = p.new_tree("t")
        src = t.add_element(Source(v_min=5, v_typ=5, v_max=5,
                                   limit_type="current", limit_value=2.0))
        t.add_element(Load(load_type=LoadType.CURRENT, value_typ=1.9),
                      parent_id=src.id)    # 95 % of limit -> warning only
        api.save(p, clean)
        normal = subprocess.run(
            [sys.executable, "-m", "powertree", "validate", clean],
            capture_output=True, text=True, env=env, cwd=root, timeout=120)
        strict = subprocess.run(
            [sys.executable, "-m", "powertree", "validate", clean,
             "--strict"],
            capture_output=True, text=True, env=env, cwd=root, timeout=120)
    assert normal.returncode == 0, normal.stdout + normal.stderr
    assert strict.returncode == 1, strict.stdout


def test_growth_analysis_hand_checkable():
    from powertree.model.elements import LimitType
    t = PowerTree("t")
    src = t.add_element(Source(v_min=5, v_typ=5, v_max=5,
                               limit_type=LimitType.CURRENT,
                               limit_value=2.0))
    t.add_element(Load(load_type=LoadType.CURRENT, value_typ=1.0),
                  parent_id=src.id)
    g = api.growth_analysis(t)
    # limit 2 A, load 1 A -> exactly +100 % growth capacity
    assert abs(g["max_growth_pct"] - 100.0) < 2.0
    assert g["bottleneck"] is not None


def test_growth_analysis_demo_respects_waivers():
    from powertree.model.calc import solve_tree as _solve
    p = api.demo_project()
    tree = p.trees[0]
    g0 = api.growth_analysis(tree, p)
    assert g0["max_growth_pct"] == 0.0        # VDDA violation at nominal
    # waive the VDDA findings -> growth is limited by the core buck instead
    r = _solve(tree)
    for w in [w for w in r.warnings if "VDDA" in w.message]:
        api.waive_finding(p, w.element_id, w.message, "vendor errata")
    g1 = api.growth_analysis(tree, p)
    assert g1["max_growth_pct"] > 0.0
    assert g1["bottleneck"] is not None


def test_user_templates_from_json(tmp_path, monkeypatch):
    import json
    from powertree import templates as T
    payload = [{
        "key": "my_fpga", "name": "My Custom FPGA", "category": "User",
        "part_number": "XCU55C", "rails": ["0.85V"],
        "items": [{"kind": "load", "name": "VCCINT_U",
                   "rail": "0.85V",
                   "params": {"load_type": "current", "value_typ": 5.0,
                              "v_in_min": 0.83, "v_in_max": 0.87}}]}]
    path = tmp_path / "templates.json"
    path.write_text(json.dumps(payload), encoding="utf-8")
    monkeypatch.setenv("POWERTREE_TEMPLATES", str(path))
    keys = [t.key for t in T.all_templates()]
    assert "my_fpga" in keys and "zynq7020" in keys
    tpl = T.template_by_key("my_fpga")
    assert tpl.part_number == "XCU55C"
    # instantiable like a built-in
    tree = PowerTree("t")
    src = tree.add_element(Source(v_min=0.85, v_typ=0.85, v_max=0.85))
    created = T.instantiate_template(tree, tpl, {"0.85V": src.id},
                                     refdes="U9")
    assert created[0].value_typ == 5.0
    # bad file must not break anything
    bad = tmp_path / "bad.json"
    bad.write_text("{not json", encoding="utf-8")
    monkeypatch.setenv("POWERTREE_TEMPLATES", str(bad))
    assert any(t.key == "zynq7020" for t in T.all_templates())


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
