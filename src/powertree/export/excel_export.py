"""Excel report export.

Always produces a richly formatted workbook via openpyxl:
  - Overview sheet (project + per-tree summary)
  - one sheet per power tree, hierarchy indented AND grouped with Excel's
    native outline (+/- collapse), color-coded by element kind, live formulas
    (P = V x I), margin status column
  - Warnings sheet

Macro-enabled (.xlsm) export: the workbook is written, then Excel (COM) injects
a navigation/outline VBA module and saves as .xlsm. If Excel or VBA trust is
unavailable, the report falls back to .xlsx plus a PowerTree_Macros.bas file
the user can import (Alt+F11 > File > Import).
"""

from __future__ import annotations

import os
import re
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..model.elements import Project, PowerTree, ElementKind, LimitType, LoadType
from ..model.calc import solve_tree, TreeResults, block_power, fmt_si

DARK = "FF1A2030"
KIND_FILLS = {
    ElementKind.SOURCE: "FFFDF1DC",
    ElementKind.CONVERTER: "FFE2EDFD",
    ElementKind.LOAD: "FFE0F5EC",
    ElementKind.SERIES: "FFEEF1F6",
}
SEV_FILLS = {"error": "FFFDE3E8", "warn": "FFFDF3D7"}
THIN = Side(style="thin", color="FFC3CCDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

VBA_MODULE = r'''
Attribute VB_Name = "PowerTreeMacros"
' PowerTree report macros -------------------------------------------------
' ExpandAllLevels    : open every branch of the active tree sheet
' CollapseToRails    : show only source + converters (outline level 2)
' NextFinding        : jump to the next row whose Status is not OK
' HighlightOverloads : bold-red every violation row on the active sheet

Sub ExpandAllLevels()
    On Error Resume Next
    ActiveSheet.Outline.ShowLevels RowLevels:=8
End Sub

Sub CollapseToRails()
    On Error Resume Next
    ActiveSheet.Outline.ShowLevels RowLevels:=2
End Sub

Sub NextFinding()
    Dim c As Range, start As Long, r As Long, lastRow As Long
    Dim statusCol As Long
    statusCol = FindStatusColumn()
    If statusCol = 0 Then Exit Sub
    lastRow = ActiveSheet.Cells(ActiveSheet.Rows.Count, 1).End(xlUp).Row
    start = ActiveCell.Row + 1
    For r = start To lastRow
        If InStr(1, CStr(ActiveSheet.Cells(r, statusCol).Value), "OK") = 0 _
           And Len(CStr(ActiveSheet.Cells(r, statusCol).Value)) > 0 Then
            ActiveSheet.Cells(r, statusCol).Select
            Exit Sub
        End If
    Next r
    MsgBox "No further findings below the cursor.", vbInformation, "PowerTree"
End Sub

Sub HighlightOverloads()
    Dim r As Long, lastRow As Long, statusCol As Long
    statusCol = FindStatusColumn()
    If statusCol = 0 Then Exit Sub
    lastRow = ActiveSheet.Cells(ActiveSheet.Rows.Count, 1).End(xlUp).Row
    For r = 2 To lastRow
        If InStr(1, CStr(ActiveSheet.Cells(r, statusCol).Value), "VIOLATION") > 0 Then
            ActiveSheet.Rows(r).Font.Bold = True
            ActiveSheet.Rows(r).Font.Color = RGB(196, 30, 58)
        End If
    Next r
End Sub

Private Function FindStatusColumn() As Long
    Dim c As Long
    For c = 1 To 40
        If CStr(ActiveSheet.Cells(1, c).Value) = "Status" Then
            FindStatusColumn = c
            Exit Function
        End If
    Next c
    FindStatusColumn = 0
End Function
'''


def _safe_sheet_name(name: str, used: set) -> str:
    clean = re.sub(r"[\[\]:*?/\\]", "-", name)[:28] or "Tree"
    base, i = clean, 2
    while clean.lower() in used:
        clean = f"{base[:25]}_{i}"
        i += 1
    used.add(clean.lower())
    return clean


def _header(ws, row, headers, widths):
    for col, (title, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(bold=True, color="FFFFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width


def _status_text(results: TreeResults, el_id: str) -> str:
    warns = results.warnings_for(el_id)
    if not warns:
        return "OK"
    worst = results.worst_severity(el_id)
    return f"{'VIOLATION' if worst == 'error' else 'LOW MARGIN'} ({len(warns)})"


TREE_HEADERS = ["Element", "Type", "RefDes", "Signal", "Part #", "Pins",
                "Vin typ (V)", "Iin typ (A)", "Pin typ (W)",
                "Vout typ (V)", "Iout typ (A)", "Pout typ (W)", "Loss typ (W)",
                "Pin max (W)", "Load value", "Limit", "% used (max)",
                "Allowed Vin (V)", "Block", "Status", "Notes"]
TREE_WIDTHS = [32, 11, 9, 14, 14, 12, 10, 10, 10, 10, 10, 10, 10,
               10, 14, 12, 11, 14, 16, 15, 34]


def _tree_sheet(wb: Workbook, tree: PowerTree, results: TreeResults, used: set):
    ws = wb.create_sheet(_safe_sheet_name(tree.name, used))
    _header(ws, 1, TREE_HEADERS, TREE_WIDTHS)
    ws.freeze_panes = "B2"
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TREE_HEADERS))}1"

    row = [2]

    def emit(el, depth):
        r = row[0]
        typ = results.get(el.id, "typ")
        mx = results.get(el.id, "max")
        block = tree.blocks.get(el.block_id) if el.block_id else None
        if el.kind == ElementKind.LOAD:
            unit = "A" if el.load_type == LoadType.CURRENT else "W"
            peak = f" / {el.value_max:g}" if el.value_max is not None else ""
            load_val = f"{el.value_typ:g}{peak} {unit}"
        else:
            load_val = ""
        limit = ""
        pct_used = None
        limit_type = getattr(el, "limit_type", LimitType.NONE)
        limit_value = getattr(el, "limit_value", 0.0)
        if limit_type != LimitType.NONE and limit_value > 0:
            unit = "A" if limit_type == LimitType.CURRENT else "W"
            limit = f"{limit_value:g} {unit}"
            used_val = mx.i_out if limit_type == LimitType.CURRENT else mx.p_out
            if el.kind == ElementKind.SOURCE:
                used_val = mx.i_out if limit_type == LimitType.CURRENT else mx.p_out
            pct_used = used_val / limit_value * 100
        window = ""
        if getattr(el, "v_in_min", None) is not None or \
           getattr(el, "v_in_max", None) is not None:
            lo = f"{el.v_in_min:g}" if el.v_in_min is not None else "—"
            hi = f"{el.v_in_max:g}" if el.v_in_max is not None else "—"
            window = f"{lo} … {hi}"

        values = ["  " * depth + el.name, el.kind, el.refdes, el.signal_name,
                  el.part_number, el.pins,
                  round(typ.v_in, 6), round(typ.i_in, 9), None,
                  round(typ.v_out, 6) if el.kind != ElementKind.LOAD else None,
                  round(typ.i_out, 9) if el.kind != ElementKind.LOAD else None,
                  round(typ.p_out, 9) if el.kind != ElementKind.LOAD else None,
                  round(typ.p_loss, 9) if typ.p_loss > 1e-12 else None,
                  round(mx.p_in, 9), load_val, limit,
                  round(pct_used, 1) if pct_used is not None else None,
                  window, block.name if block else "",
                  _status_text(results, el.id), el.description]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = BORDER
            cell.font = Font(size=9)
            if col in (7, 8, 9, 10, 11, 12, 13, 14, 17):
                cell.number_format = "0.000###"
                cell.alignment = Alignment(horizontal="right")
        # live formula: P = V x I (holds for every element type)
        ws.cell(row=r, column=9,
                value=f"=ROUND(G{r}*H{r},6)").number_format = "0.000###"

        sev = results.worst_severity(el.id)
        fill = SEV_FILLS.get(sev) or KIND_FILLS.get(el.kind)
        if fill:
            for col in range(1, len(TREE_HEADERS) + 1):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=fill)
        if sev == "error":
            ws.cell(row=r, column=20).font = Font(size=9, bold=True,
                                                  color="FFC41E3A")
        ws.row_dimensions[r].outlineLevel = min(depth, 7)
        row[0] += 1
        for child in tree.children_of(el.id):
            emit(child, depth + 1)

    if tree.source:
        emit(tree.source, 0)

    # block summary below the tree
    if tree.blocks:
        r = row[0] + 1
        ws.cell(row=r, column=1, value="Blocks").font = Font(bold=True, size=10)
        r += 1
        _header(ws, r, ["Block", "Members", "P typ (W)", "P max (W)"],
                [32, 11, 12, 12])
        for bid, block in tree.blocks.items():
            r += 1
            vals = [block.name, len(tree.block_members(bid)),
                    round(block_power(tree, results, bid, "typ"), 6),
                    round(block_power(tree, results, bid, "max"), 6)]
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = BORDER
                cell.font = Font(size=9)
    return ws


def _overview_sheet(wb: Workbook, project: Project, tree_results: dict):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=project.name).font = Font(bold=True, size=16)
    ws.cell(row=2, column=1, value="Power Tree Analysis Report").font = \
        Font(size=11, color="FF667085")
    if project.description:
        ws.cell(row=3, column=1, value=project.description).font = Font(size=9)
    from ..api import tree_metrics
    _header(ws, 5, ["Power tree", "Elements", "Source", "P typ (W)",
                    "P max (W)", "Efficiency (%)", "Loss typ (W)",
                    "Errors", "Warnings"],
            [34, 10, 26, 12, 12, 13, 12, 9, 10])
    r = 5
    for tree in project.trees:
        results = tree_results[tree.id]
        r += 1
        src = tree.source
        typ = results.get(src.id, "typ") if src else None
        mx = results.get(src.id, "max") if src else None
        metrics = tree_metrics(tree, results)
        errs = sum(1 for w in results.warnings if w.severity == "error")
        warns = sum(1 for w in results.warnings if w.severity == "warn")
        vals = [tree.name, len(tree.elements), src.name if src else "—",
                round(typ.p_out, 6) if typ else None,
                round(mx.p_out, 6) if mx else None,
                metrics["efficiency_pct"],
                round(metrics["p_loss_typ"], 6), errs, warns]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = BORDER
            cell.font = Font(size=9)
        if errs:
            ws.cell(row=r, column=8).fill = PatternFill(
                "solid", fgColor=SEV_FILLS["error"])
        if warns:
            ws.cell(row=r, column=9).fill = PatternFill(
                "solid", fgColor=SEV_FILLS["warn"])
    r += 2
    ws.cell(row=r, column=1,
            value="Each tree has its own sheet. Use the +/- outline buttons on "
                  "the left edge to collapse or expand branches; with macros "
                  "enabled, run PowerTreeMacros (Alt+F8) for one-click "
                  "expand / collapse / next-finding navigation.").font = \
        Font(size=8, italic=True, color="FF667085")


def _warnings_sheet(wb: Workbook, project: Project, tree_results: dict):
    ws = wb.create_sheet("Warnings")
    _header(ws, 1, ["Power tree", "Severity", "Corner", "Element", "Message"],
            [24, 10, 8, 24, 90])
    ws.freeze_panes = "A2"
    r = 1
    for tree in project.trees:
        results = tree_results[tree.id]
        for w in results.warnings:
            r += 1
            el = tree.elements.get(w.element_id) if w.element_id else None
            vals = [tree.name, w.severity.upper(), w.corner,
                    el.name if el else "—", w.message]
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = BORDER
                cell.font = Font(size=9)
            fill = SEV_FILLS.get(w.severity)
            if fill:
                for col in range(1, 6):
                    ws.cell(row=r, column=col).fill = PatternFill(
                        "solid", fgColor=fill)
    if r == 1:
        ws.cell(row=2, column=1, value="No findings — all margins healthy.")


def _states_sheet(wb: Workbook, project: Project):
    """Per-operating-state comparison: source power + findings per tree."""
    ws = wb.create_sheet("States")
    _header(ws, 1, ["Power tree", "State", "P min (W)", "P typ (W)",
                    "P max (W)", "Errors", "Warnings"],
            [28, 18, 12, 12, 12, 9, 10])
    ws.freeze_panes = "A2"
    r = 1
    for tree in project.trees:
        src = tree.source
        if src is None:
            continue
        for label, scenario in [("Base", None)] + \
                [(s, s) for s in project.scenarios]:
            sr = solve_tree(tree, scenario)
            errs = sum(1 for w in sr.warnings if w.severity == "error")
            warns = sum(1 for w in sr.warnings if w.severity == "warn")
            r += 1
            vals = [tree.name, label,
                    round(sr.get(src.id, "min").p_out, 6),
                    round(sr.get(src.id, "typ").p_out, 6),
                    round(sr.get(src.id, "max").p_out, 6), errs, warns]
            for col, val in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = BORDER
                cell.font = Font(size=9)
            if errs:
                ws.cell(row=r, column=6).fill = PatternFill(
                    "solid", fgColor=SEV_FILLS["error"])
            if warns:
                ws.cell(row=r, column=7).fill = PatternFill(
                    "solid", fgColor=SEV_FILLS["warn"])


def export_excel_xlsx(project: Project, path: str) -> str:
    wb = Workbook()
    tree_results = {t.id: solve_tree(t) for t in project.trees}
    _overview_sheet(wb, project, tree_results)
    used: set = {"overview", "warnings", "states"}
    for tree in project.trees:
        _tree_sheet(wb, tree, tree_results[tree.id], used)
    if project.scenarios:
        _states_sheet(wb, project)
    _warnings_sheet(wb, project, tree_results)
    wb.save(path)
    return path


def export_excel_xlsm(project: Project, path: str) -> tuple:
    """Write a macro-enabled report. Returns (written_path, info_message).

    Requires Excel + 'Trust access to the VBA project object model'. Falls back
    to .xlsx alongside a PowerTree_Macros.bas import file when unavailable.
    """
    tmp_xlsx = os.path.join(tempfile.gettempdir(), "_powertree_report_tmp.xlsx")
    export_excel_xlsx(project, tmp_xlsx)
    xlsm_abs = os.path.abspath(path)
    try:
        import win32com.client
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(os.path.abspath(tmp_xlsx))
            module = wb.VBProject.VBComponents.Add(1)   # vbext_ct_StdModule
            module.CodeModule.AddFromString(VBA_MODULE)
            if os.path.exists(xlsm_abs):
                os.remove(xlsm_abs)
            wb.SaveAs(xlsm_abs, FileFormat=52)          # xlOpenXMLWorkbookMacroEnabled
            wb.Close(False)
            return xlsm_abs, "Macro-enabled report written (VBA module embedded)."
        finally:
            excel.Quit()
    except Exception as exc:
        fallback = os.path.splitext(path)[0] + ".xlsx"
        export_excel_xlsx(project, fallback)
        bas_path = os.path.join(os.path.dirname(xlsm_abs) or ".",
                                "PowerTree_Macros.bas")
        with open(bas_path, "w", encoding="utf-8") as fh:
            fh.write(VBA_MODULE)
        return fallback, (
            "Excel VBA injection unavailable "
            f"({type(exc).__name__}: {exc}). Wrote {os.path.basename(fallback)} "
            "plus PowerTree_Macros.bas — import it via Alt+F11 > File > Import "
            "and save as .xlsm, or enable File > Options > Trust Center > "
            "'Trust access to the VBA project object model' and re-export.")
    finally:
        try:
            os.remove(tmp_xlsx)
        except OSError:
            pass
